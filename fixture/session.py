# -*- coding: utf-8 -*-
from openpyxl import load_workbook


class SessionHelper():

    def __init__(self, app):
        self.app = app

    def logout(self):
        wd = self.app.wd
        wd.find_element_by_link_text("Вийти").click()

    def is_logged_in(self):
        wd = self.app.wd
        return len(wd.find_elements_by_link_text("Вийти")) > 0

    def is_logged_in_as(self, username):
        wd = self.app.wd
        return self.get_logged_user() == username

    def get_logged_user(self):
        wd = self.app.wd
        # и избавляемся от обрамляющих его круглых скобок - (username)
        return wd.find_element_by_xpath("//div[@id='top']/form[@class='header']/b").text[1:-1]

    def ensure_logout(self):
        wd = self.app.wd
        if self.is_logged_in():
            self.logout()

    def ensure_login(self, username, password):
        wd = self.app.wd
        if self.is_logged_in():
            if self.is_logged_in_as(username):
                return
            else:
                self.logout()
        self.login(username, password)

    def login_from_excel(self):
        wd = self.app.wd
        self.app.open_home_page()
        wb = load_workbook(filename="C:\\Reposit\\soft_sob\\book1.xlsx")
        ws = wb['Sheet1']
        row_number = 1
        for row in ws.rows:
            wd.get('https://www.addressbook.net/cgi-bin/WebObjects/AddressBook.woa/wa/UserAction/login')
            wd.find_element_by_id('login').send_keys(row[0].value)
            wd.find_element_by_id('password').send_keys(row[1].value)
            wd.find_element_by_id('loginSubmit').click()

            ws['C' + str(row_number)] = "OK"
            wb.save(filename="C:\\Reposit\\soft_sob\\data_result.xlsx")
            row_number += 1
        else:
            pass